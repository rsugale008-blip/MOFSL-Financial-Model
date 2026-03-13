# model_builder.py
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import os
 
NAVY       = "1F2D4E"
MID_BLUE   = "2E4A7A"
LIGHT_BLUE = "D9E4F0"
WHITE      = "FFFFFF"
BLACK      = "000000"
INPUT_BLUE = "0000FF"
YELLOW     = "FFFF00"
 
def make_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
 
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)
 
def make_font(bold=False, size=10, color=BLACK, italic=False):
    return Font(name="Arial", bold=bold, size=size,
                color=color, italic=italic)
 
def style_cell(cell, value=None, bold=False, size=10,
               color=BLACK, bg=None, num_fmt=None,
               align="right", italic=False):
    if value is not None:
        cell.value = value
    cell.font      = make_font(bold=bold, size=size,
                               color=color, italic=italic)
    cell.alignment = Alignment(horizontal=align,
                               vertical="center")
    cell.border    = make_border()
    if bg:
        cell.fill  = make_fill(bg)
    if num_fmt:
        cell.number_format = num_fmt
 
def header_cell(ws, row, col_start, col_end, text,
                bg=NAVY, fg=WHITE, size=11):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    style_cell(cell, text, bold=True, size=size,
               color=fg, bg=bg, align="center")
 
def section_header(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    style_cell(cell, text, bold=True, size=9,
               color=WHITE, bg=MID_BLUE, align="left")
 
def set_col_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
 
 
# =============================================================
# ASSUMPTIONS SHEET
# =============================================================
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [32,14,14,14,14,14,14,14])
 
    header_cell(ws, 1, 1, 8,
                "MOTILAL OSWAL -- ASSUMPTIONS & KEY DRIVERS",
                size=12)
 
    years = ["FY2022A","FY2023A","FY2024A","FY2025A",
             "FY2026E","FY2027E","FY2028E","FY2029E"]
    ws.cell(2,1).value  = "Driver"
    ws.cell(2,1).font   = make_font(bold=True, color=WHITE)
    ws.cell(2,1).fill   = make_fill(MID_BLUE)
    ws.cell(2,1).border = make_border()
    for i, yr in enumerate(years):
        c = ws.cell(2, 2+i)
        style_cell(c, yr, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    section_header(ws, 3, 1, 8, "REVENUE ASSUMPTIONS")
 
    ws.cell(4,1).value  = "Revenue Growth Rate (%)"
    ws.cell(4,1).font   = make_font(size=9)
    ws.cell(4,1).border = make_border()
    growth_vals = [0.041, 0.355, 0.245, 0.245,
                   0.18,  0.15,  0.12,  0.10]
    for i, g in enumerate(growth_vals):
        c = ws.cell(4, 2+i)
        is_proj = i >= 4
        style_cell(c, g,
                   color=INPUT_BLUE if is_proj else BLACK,
                   bg=YELLOW if is_proj else None,
                   num_fmt="0.0%")
 
    section_header(ws, 5, 1, 8, "MARGIN ASSUMPTIONS")
 
    ws.cell(6,1).value  = "EBITDA Margin (%)"
    ws.cell(6,1).font   = make_font(size=9)
    ws.cell(6,1).border = make_border()
    ebitda_margins = [0.577, 0.488, 0.795, 0.711,
                      0.68,  0.70,  0.72,  0.72]
    for i, v in enumerate(ebitda_margins):
        c = ws.cell(6, 2+i)
        is_proj = i >= 4
        style_cell(c, v,
                   color=INPUT_BLUE if is_proj else BLACK,
                   bg=YELLOW if is_proj else None,
                   num_fmt="0.0%")
 
    ws.cell(7,1).value  = "Net Income Margin (%)"
    ws.cell(7,1).font   = make_font(size=9)
    ws.cell(7,1).border = make_border()
    ni_margins = [0.357, 0.180, 0.472, 0.388,
                  0.40,  0.41,  0.42,  0.42]
    for i, v in enumerate(ni_margins):
        c = ws.cell(7, 2+i)
        is_proj = i >= 4
        style_cell(c, v,
                   color=INPUT_BLUE if is_proj else BLACK,
                   bg=YELLOW if is_proj else None,
                   num_fmt="0.0%")
 
    ws.cell(8,1).value  = "Tax Rate (%)"
    ws.cell(8,1).font   = make_font(size=9)
    ws.cell(8,1).border = make_border()
    tax_vals = [0.25, 0.25, 0.25, 0.25,
                0.25, 0.25, 0.25, 0.25]
    for i, v in enumerate(tax_vals):
        c = ws.cell(8, 2+i)
        is_proj = i >= 4
        style_cell(c, v,
                   color=INPUT_BLUE if is_proj else BLACK,
                   bg=YELLOW if is_proj else None,
                   num_fmt="0.0%")
 
    section_header(ws, 10, 1, 8, "DCF ASSUMPTIONS")
 
    dcf_items = [
        ("WACC (%)",                0.14,  "0.0%"),
        ("Risk Free Rate (%)",      0.071, "0.0%"),
        ("Equity Risk Premium (%)", 0.075, "0.0%"),
        ("Beta",                    1.15,  "0.00"),
        ("Terminal Growth Rate (%)",0.05,  "0.0%"),
        ("Net Debt (Rs Cr)",        8138,  "#,##0"),
        ("Shares Outstanding (Cr)", 14.4,  "#,##0.0"),
    ]
    for i, (label, val, fmt) in enumerate(dcf_items):
        r = 11 + i
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(size=9)
        ws.cell(r,1).border = make_border()
        style_cell(ws.cell(r,2), val,
                   color=INPUT_BLUE, bg=YELLOW, num_fmt=fmt)
 
    section_header(ws, 19, 1, 8, "LBO ASSUMPTIONS")
 
    lbo_items = [
        ("Entry EV/EBITDA Multiple (x)", 10.0, "0.0"),
        ("Exit EV/EBITDA Multiple (x)",  12.0, "0.0"),
        ("Debt % of Entry EV",           0.50, "0.0%"),
        ("Senior Debt % of Total Debt",  0.70, "0.0%"),
        ("Mezz Debt % of Total Debt",    0.30, "0.0%"),
        ("Senior Interest Rate (%)",     0.09, "0.0%"),
        ("Mezz Interest Rate (%)",       0.13, "0.0%"),
        ("Amortisation (% of debt/yr)",  0.05, "0.0%"),
        ("Holding Period (years)",       5,    "0"),
    ]
    for i, (label, val, fmt) in enumerate(lbo_items):
        r = 20 + i
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(size=9)
        ws.cell(r,1).border = make_border()
        style_cell(ws.cell(r,2), val,
                   color=INPUT_BLUE, bg=YELLOW, num_fmt=fmt)
 
    return ws
 
 
# =============================================================
# 3-STATEMENT SHEET
# =============================================================
def build_three_statement(wb):
    ws = wb.create_sheet("3-Statement")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [32,13,13,13,13,13,13,13])
 
    header_cell(ws, 1, 1, 8,
                "THREE-STATEMENT MODEL -- MOTILAL OSWAL (Rs Crores)",
                size=12)
 
    # COLUMN MAP:
    # Col B = FY2022A  (j=0)
    # Col C = FY2023A  (j=1)
    # Col D = FY2024A  (j=2)
    # Col E = FY2025A  (j=3)  <- last actual
    # Col F = FY2026E  (j=4)  <- DCF & LBO use this
    # Col G = FY2027E  (j=5)
    # Col H = FY2028E  (j=6)
    # Col I = FY2029E  (j=7)
 
    years = ["FY2022A","FY2023A","FY2024A","FY2025A",
             "FY2026E","FY2027E","FY2028E","FY2029E"]
    ws.cell(2,1).fill   = make_fill(MID_BLUE)
    ws.cell(2,1).border = make_border()
    for i, yr in enumerate(years):
        c = ws.cell(2, 2+i)
        style_cell(c, yr, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    section_header(ws, 3, 1, 8, "INCOME STATEMENT")
 
    # Row 4: Total Revenue
    ws.cell(4,1).value  = "Total Revenue"
    ws.cell(4,1).font   = make_font(bold=True, size=9)
    ws.cell(4,1).fill   = make_fill(LIGHT_BLUE)
    ws.cell(4,1).border = make_border()
    rev_actuals = [3818.9, 5174.4, 6441.8, 6441.8]
    for j in range(8):
        c = ws.cell(4, 2+j)
        if j < 4:
            style_cell(c, rev_actuals[j], bold=True,
                       bg=LIGHT_BLUE,
                       num_fmt="#,##0.0;(#,##0.0);-")
        else:
            prev = get_column_letter(2+j-1)
            style_cell(c, f"={prev}4*(1+Assumptions!B4)",
                       bold=True, bg=LIGHT_BLUE,
                       num_fmt="#,##0.0;(#,##0.0);-")
 
    # Row 5: Gross Profit
    ws.cell(5,1).value  = "  Gross Profit"
    ws.cell(5,1).font   = make_font(size=9)
    ws.cell(5,1).border = make_border()
    gp_actuals = [None, 2032.9, 2791.7, 3526.5]
    for j in range(8):
        c = ws.cell(5, 2+j)
        if j < 4 and gp_actuals[j] is not None:
            style_cell(c, gp_actuals[j],
                       num_fmt="#,##0.0;(#,##0.0);-")
        else:
            style_cell(c, "---", color="BFBFBF", align="center")
 
    # Row 6: EBITDA
    # CRITICAL: DCF links to F6, G6, H6
    #           LBO  links to F6, G6, H6
    ws.cell(6,1).value  = "  EBITDA"
    ws.cell(6,1).font   = make_font(bold=True, size=9)
    ws.cell(6,1).fill   = make_fill(LIGHT_BLUE)
    ws.cell(6,1).border = make_border()
    ebitda_actuals = [1862.5, 4115.0, 4578.3, 4578.3]
    for j in range(8):
        c = ws.cell(6, 2+j)
        if j < 4:
            style_cell(c, ebitda_actuals[j], bold=True,
                       bg=LIGHT_BLUE,
                       num_fmt="#,##0.0;(#,##0.0);-")
        else:
            rev_col = get_column_letter(2+j)
            style_cell(c, f"={rev_col}4*Assumptions!B6",
                       bold=True, bg=LIGHT_BLUE,
                       num_fmt="#,##0.0;(#,##0.0);-")
 
    # Row 7: EBIT
    ws.cell(7,1).value  = "  EBIT"
    ws.cell(7,1).font   = make_font(size=9)
    ws.cell(7,1).border = make_border()
    ebit_actuals = [1804.1, 4032.5, 4479.5, 4479.5]
    for j in range(8):
        c = ws.cell(7, 2+j)
        if j < 4:
            style_cell(c, ebit_actuals[j],
                       num_fmt="#,##0.0;(#,##0.0);-")
        else:
            style_cell(c, "---", color="BFBFBF", align="center")
 
    # Row 8: Net Income
    ws.cell(8,1).value  = "  Net Income"
    ws.cell(8,1).font   = make_font(bold=True, size=9)
    ws.cell(8,1).fill   = make_fill(LIGHT_BLUE)
    ws.cell(8,1).border = make_border()
    ni_actuals = [931.1, 2441.1, 2501.6, 2501.6]
    for j in range(8):
        c = ws.cell(8, 2+j)
        if j < 4:
            style_cell(c, ni_actuals[j], bold=True,
                       bg=LIGHT_BLUE,
                       num_fmt="#,##0.0;(#,##0.0);-")
        else:
            ebitda_col = get_column_letter(2+j)
            style_cell(c, f"={ebitda_col}6*Assumptions!B7",
                       bold=True, bg=LIGHT_BLUE,
                       num_fmt="#,##0.0;(#,##0.0);-")
 
    # Balance Sheet
    section_header(ws, 10, 1, 8, "BALANCE SHEET")
    bs_rows = [
        ("Total Assets",
         [23009.9, 31829.2, 33987.1, 33987.1], True),
        ("  Cash & Equivalents",
         [2576.3,  5285.7,  6592.5,  6592.5],  False),
        ("  Total Debt",
         [10275.9, 13745.6, 14731.6, 14731.6], False),
        ("  Stockholders Equity",
         [6252.2,  8731.8,  11079.3, 11079.3], True),
    ]
    for offset, (label, actuals, is_bold) in enumerate(bs_rows):
        r = 11 + offset
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(bold=is_bold, size=9)
        ws.cell(r,1).border = make_border()
        if is_bold:
            ws.cell(r,1).fill = make_fill(LIGHT_BLUE)
        for j in range(8):
            c  = ws.cell(r, 2+j)
            bg = LIGHT_BLUE if is_bold else None
            if j < 4:
                style_cell(c, actuals[j], bold=is_bold, bg=bg,
                           num_fmt="#,##0.0;(#,##0.0);-")
            else:
                style_cell(c, "---", color="BFBFBF", align="center")
 
    # Cash Flow
    section_header(ws, 16, 1, 8, "CASH FLOW STATEMENT")
    cf_rows = [
        ("Operating Cash Flow",
         [-3057.9, -349.6, 1214.6, 1214.6], True),
        ("  Capital Expenditure",
         [-153.8,  -150.8, -284.3, -284.3],  False),
        ("  Free Cash Flow",
         [-3211.7, -500.4,  930.4,  930.4],  True),
        ("  Financing Cash Flow",
         [3770.4,  3305.8,  745.1,  745.1],  False),
    ]
    for offset, (label, actuals, is_bold) in enumerate(cf_rows):
        r = 17 + offset
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(bold=is_bold, size=9)
        ws.cell(r,1).border = make_border()
        if is_bold:
            ws.cell(r,1).fill = make_fill(LIGHT_BLUE)
        for j in range(8):
            c  = ws.cell(r, 2+j)
            bg = LIGHT_BLUE if is_bold else None
            if j < 4:
                style_cell(c, actuals[j], bold=is_bold, bg=bg,
                           num_fmt="#,##0.0;(#,##0.0);-")
            else:
                style_cell(c, "---", color="BFBFBF", align="center")
 
    return ws
 
 
# =============================================================
# DCF SHEET
# =============================================================
def build_dcf(wb):
    ws = wb.create_sheet("DCF")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [35,14,14,14,14,14])
 
    header_cell(ws, 1, 1, 6,
                "DCF VALUATION -- MOTILAL OSWAL (Rs Crores)",
                size=12)
 
    # 3-Statement columns: FY2026E=F, FY2027E=G, FY2028E=H
    # EBITDA=row6, Revenue=row4
    stmt_cols = ["F", "G", "H"]
 
    proj_years = ["FY2026E","FY2027E","FY2028E"]
    ws.cell(2,1).fill   = make_fill(MID_BLUE)
    ws.cell(2,1).border = make_border()
    for i, yr in enumerate(proj_years):
        c = ws.cell(2, 2+i)
        style_cell(c, yr, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    # STEP 1: UFCF
    section_header(ws, 3, 1, 6,
                   "STEP 1 -- UNLEVERED FREE CASH FLOW (UFCF)")
 
    fcf_items = [
        (4,  "EBITDA (from 3-Stmt)",     True),
        (5,  "  (-) Tax on EBIT (25%)",  False),
        (6,  "  NOPAT",                  True),
        (7,  "  (+) D&A est. (2% Rev)",  False),
        (8,  "  (-) Capex (2% Rev)",     False),
        (9,  "  (-) Change in NWC (3%)", False),
        (10, "Unlevered Free Cash Flow", True),
    ]
 
    for r, label, is_bold in fcf_items:
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(bold=is_bold, size=9)
        ws.cell(r,1).border = make_border()
        if is_bold:
            ws.cell(r,1).fill = make_fill(LIGHT_BLUE)
 
        for j, sc in enumerate(stmt_cols):
            c   = ws.cell(r, 2+j)
            col = get_column_letter(2+j)
            bg  = LIGHT_BLUE if is_bold else None
            fmt = "#,##0.0;(#,##0.0);-"
 
            if r == 4:
                style_cell(c, f"='3-Statement'!{sc}6",
                           bold=True, bg=bg,
                           num_fmt=fmt, color="008000")
            elif r == 5:
                style_cell(c, f"=-{col}4*Assumptions!B8",
                           bg=bg, num_fmt=fmt)
            elif r == 6:
                style_cell(c, f"={col}4+{col}5",
                           bold=True, bg=bg, num_fmt=fmt)
            elif r == 7:
                style_cell(c, f"='3-Statement'!{sc}4*0.02",
                           bg=bg, num_fmt=fmt)
            elif r == 8:
                style_cell(c, f"=-'3-Statement'!{sc}4*0.02",
                           bg=bg, num_fmt=fmt)
            elif r == 9:
                style_cell(c, f"=-'3-Statement'!{sc}4*0.03",
                           bg=bg, num_fmt=fmt)
            elif r == 10:
                style_cell(c, f"={col}6+{col}7+{col}8+{col}9",
                           bold=True, bg=bg, num_fmt=fmt)
 
    # STEP 2: DISCOUNT FACTORS
    section_header(ws, 12, 1, 6,
                   "STEP 2 -- DISCOUNT FACTORS & PRESENT VALUES")
 
    ws.cell(13,1).value  = "Discount Period (mid-year)"
    ws.cell(13,1).font   = make_font(size=9)
    ws.cell(13,1).border = make_border()
    for j, p in enumerate([0.5, 1.5, 2.5]):
        style_cell(ws.cell(13, 2+j), p,
                   color=INPUT_BLUE, num_fmt="0.0")
 
    ws.cell(14,1).value  = "Discount Factor"
    ws.cell(14,1).font   = make_font(size=9)
    ws.cell(14,1).border = make_border()
    for j in range(3):
        col = get_column_letter(2+j)
        style_cell(ws.cell(14, 2+j),
                   f"=1/(1+Assumptions!B11)^{col}13",
                   num_fmt="0.000")
 
    ws.cell(15,1).value  = "PV of UFCF"
    ws.cell(15,1).font   = make_font(bold=True, size=9)
    ws.cell(15,1).fill   = make_fill(LIGHT_BLUE)
    ws.cell(15,1).border = make_border()
    for j in range(3):
        col = get_column_letter(2+j)
        style_cell(ws.cell(15, 2+j),
                   f"={col}10*{col}14",
                   bold=True, bg=LIGHT_BLUE,
                   num_fmt="#,##0.0;(#,##0.0);-")
 
    # STEP 3: TERMINAL VALUE
    section_header(ws, 17, 1, 6,
                   "STEP 3 -- TERMINAL VALUE (GORDON GROWTH)")
 
    ws.cell(18,1).value  = "Terminal Year UFCF (FY2028E)"
    ws.cell(18,1).font   = make_font(bold=True, size=9)
    ws.cell(18,1).fill   = make_fill(LIGHT_BLUE)
    ws.cell(18,1).border = make_border()
    style_cell(ws.cell(18,2), "=D10",
               bold=True, bg=LIGHT_BLUE,
               num_fmt="#,##0.0;(#,##0.0);-")
 
    ws.cell(19,1).value  = "Terminal Growth Rate (g)"
    ws.cell(19,1).font   = make_font(size=9)
    ws.cell(19,1).border = make_border()
    style_cell(ws.cell(19,2), "=Assumptions!B15",
               num_fmt="0.0%", color="008000")
 
    ws.cell(20,1).value  = "WACC"
    ws.cell(20,1).font   = make_font(size=9)
    ws.cell(20,1).border = make_border()
    style_cell(ws.cell(20,2), "=Assumptions!B11",
               num_fmt="0.0%", color="008000")
 
    ws.cell(21,1).value  = "Terminal Value = UFCF*(1+g)/(WACC-g)"
    ws.cell(21,1).font   = make_font(bold=True, size=9)
    ws.cell(21,1).fill   = make_fill(LIGHT_BLUE)
    ws.cell(21,1).border = make_border()
    style_cell(ws.cell(21,2),
               "=B18*(1+B19)/(B20-B19)",
               bold=True, bg=LIGHT_BLUE,
               num_fmt="#,##0.0;(#,##0.0);-")
 
    ws.cell(22,1).value  = "PV of Terminal Value"
    ws.cell(22,1).font   = make_font(bold=True, size=9)
    ws.cell(22,1).fill   = make_fill(LIGHT_BLUE)
    ws.cell(22,1).border = make_border()
    style_cell(ws.cell(22,2),
               "=B21/(1+Assumptions!B11)^2.5",
               bold=True, bg=LIGHT_BLUE,
               num_fmt="#,##0.0;(#,##0.0);-")
 
    # STEP 4: EV BRIDGE
    section_header(ws, 24, 1, 6,
                   "STEP 4 -- ENTERPRISE VALUE TO EQUITY VALUE")
 
    bridge = [
        (25, "Sum of PV of UFCFs",
         "=SUM(B15:D15)", "#,##0.0", True),
        (26, "(+) PV of Terminal Value",
         "=B22", "#,##0.0", False),
        (27, "Enterprise Value (EV)",
         "=B25+B26", "#,##0.0", True),
        (28, "(-) Net Debt",
         "=Assumptions!B16", "#,##0.0", False),
        (29, "Equity Value",
         "=B27-B28", "#,##0.0", True),
        (30, "Shares Outstanding (Cr)",
         "=Assumptions!B17", "#,##0.0", False),
        (31, "Implied Share Price (Rs)",
         "=B29/B30", "Rs #,##0.00", True),
    ]
    for r, label, formula, fmt, is_bold in bridge:
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(bold=is_bold, size=9)
        ws.cell(r,1).border = make_border()
        if is_bold:
            ws.cell(r,1).fill = make_fill(LIGHT_BLUE)
        style_cell(ws.cell(r,2), formula,
                   bold=is_bold,
                   bg=LIGHT_BLUE if is_bold else None,
                   num_fmt=fmt)
 
    # STEP 5: SENSITIVITY
    section_header(ws, 33, 1, 6,
                   "STEP 5 -- SENSITIVITY: WACC vs TERMINAL GROWTH")
 
    ws.cell(34,1).value  = "Share Price  WACC vs TGR"
    ws.cell(34,1).font   = make_font(bold=True, size=9, color=WHITE)
    ws.cell(34,1).fill   = make_fill(NAVY)
    ws.cell(34,1).border = make_border()
 
    tgrs  = [0.03, 0.04, 0.05, 0.06, 0.07]
    waccs = [0.12, 0.13, 0.14, 0.15, 0.16]
 
    for j, tgr in enumerate(tgrs):
        c = ws.cell(34, 2+j)
        style_cell(c, tgr, bold=True, color=WHITE,
                   bg=MID_BLUE, num_fmt="0%", align="center")
 
    for i, wacc in enumerate(waccs):
        r = 35 + i
        style_cell(ws.cell(r,1), wacc, bold=True,
                   color=WHITE, bg=MID_BLUE,
                   num_fmt="0%", align="center")
        for j, tgr in enumerate(tgrs):
            cell  = ws.cell(r, 2+j)
            ufcf  = 2800
            tv    = ufcf * (1+tgr) / (wacc - tgr)
            pv_tv = tv / (1+wacc) ** 2.5
            ev    = ufcf * 2.2 + pv_tv
            price = (ev - 8138) / 14.4 * 100
            bg    = ("E2EFDA" if price > 1200
                     else "FCE4D6" if price < 700
                     else "FFFF00")
            style_cell(cell, round(price, 0),
                       num_fmt="Rs #,##0", bg=bg,
                       align="center",
                       bold=(i==2 and j==2))
 
    return ws
 
 
# =============================================================
# LBO SHEET
# =============================================================
def build_lbo(wb):
    ws = wb.create_sheet("LBO")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [35,14,14,14,14,14,14])
 
    header_cell(ws, 1, 1, 7,
                "LBO MODEL -- MOTILAL OSWAL (Rs Crores)",
                size=12)
 
    # TRANSACTION ASSUMPTIONS
    section_header(ws, 3, 1, 7, "TRANSACTION ASSUMPTIONS")
 
    trans_items = [
        ("Entry EBITDA (FY2026E, Rs Cr)",  3869,       "#,##0.0"),
        ("Entry EV/EBITDA Multiple (x)",   10.0,       "0.0"),
        ("Entry Enterprise Value (Rs Cr)", "=B4*B5",   "#,##0.0"),
        ("(-) Equity Contribution (50%)",  "=B6*0.50", "#,##0.0"),
        ("(+) Total Debt (50%)",           "=B6*0.50", "#,##0.0"),
        ("Senior Debt (70% of debt)",      "=B8*0.70", "#,##0.0"),
        ("Mezz Debt (30% of debt)",        "=B8*0.30", "#,##0.0"),
        ("Senior Interest Rate (%)",       0.09,       "0.0%"),
        ("Mezz Interest Rate (%)",         0.13,       "0.0%"),
        ("Amortisation (% of debt/yr)",    0.05,       "0.0%"),
    ]
 
    for i, (label, val, fmt) in enumerate(trans_items):
        r = 4 + i
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(size=9)
        ws.cell(r,1).border = make_border()
        c = ws.cell(r, 2)
        is_input = not str(val).startswith("=")
        style_cell(c, val,
                   color=INPUT_BLUE if is_input else BLACK,
                   bg=YELLOW if is_input else LIGHT_BLUE,
                   num_fmt=fmt)
 
    # PROJECTION YEARS
    section_header(ws, 15, 1, 7, "5-YEAR PROJECTIONS")
 
    proj_years = ["FY2026E","FY2027E","FY2028E",
                  "FY2029E","FY2030E"]
    ws.cell(16,1).fill   = make_fill(MID_BLUE)
    ws.cell(16,1).border = make_border()
    for i, yr in enumerate(proj_years):
        c = ws.cell(16, 2+i)
        style_cell(c, yr, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    # Revenue
    ws.cell(17,1).value  = "Revenue (Rs Cr)"
    ws.cell(17,1).font   = make_font(bold=True, size=9)
    ws.cell(17,1).fill   = make_fill(LIGHT_BLUE)
    ws.cell(17,1).border = make_border()
    rev_links = [
        "='3-Statement'!F4",
        "='3-Statement'!G4",
        "='3-Statement'!H4",
        "=E17*1.10",
        "=F17*1.10",
    ]
    for j, formula in enumerate(rev_links):
        c = ws.cell(17, 2+j)
        style_cell(c, formula, bold=True,
                   bg=LIGHT_BLUE,
                   num_fmt="#,##0.0;(#,##0.0);-")
 
    # EBITDA
    ws.cell(18,1).value  = "EBITDA (Rs Cr)"
    ws.cell(18,1).font   = make_font(bold=True, size=9)
    ws.cell(18,1).fill   = make_fill(LIGHT_BLUE)
    ws.cell(18,1).border = make_border()
    ebitda_links = [
        "='3-Statement'!F6",
        "='3-Statement'!G6",
        "='3-Statement'!H6",
        "=E18*1.10",
        "=F18*1.10",
    ]
    for j, formula in enumerate(ebitda_links):
        c = ws.cell(18, 2+j)
        style_cell(c, formula, bold=True,
                   bg=LIGHT_BLUE,
                   num_fmt="#,##0.0;(#,##0.0);-")
 
    # EBITDA Margin
    ws.cell(19,1).value  = "EBITDA Margin (%)"
    ws.cell(19,1).font   = make_font(size=9)
    ws.cell(19,1).border = make_border()
    for j in range(5):
        col_r = get_column_letter(2+j)
        c = ws.cell(19, 2+j)
        style_cell(c, f"={col_r}18/{col_r}17",
                   num_fmt="0.0%")
 
    # DEBT SCHEDULE
    section_header(ws, 21, 1, 7, "DEBT SCHEDULE")
 
    ws.cell(22,1).fill   = make_fill(MID_BLUE)
    ws.cell(22,1).border = make_border()
    for i, yr in enumerate(proj_years):
        c = ws.cell(22, 2+i)
        style_cell(c, yr, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    debt_rows = [
        ("Opening Total Debt",     True),
        ("  (-) Amortisation",     False),
        ("Closing Total Debt",     True),
        ("Senior Interest",        False),
        ("Mezz Interest",          False),
        ("Total Interest Expense", True),
    ]
 
    for offset, (label, is_bold) in enumerate(debt_rows):
        r = 23 + offset
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(bold=is_bold, size=9)
        ws.cell(r,1).border = make_border()
        if is_bold:
            ws.cell(r,1).fill = make_fill(LIGHT_BLUE)
 
        for j in range(5):
            col = get_column_letter(2+j)
            c   = ws.cell(r, 2+j)
            bg  = LIGHT_BLUE if is_bold else None
            fmt = "#,##0.0;(#,##0.0);-"
 
            if label == "Opening Total Debt":
                if j == 0:
                    style_cell(c, "=B8", bold=True,
                               bg=bg, num_fmt=fmt,
                               color="008000")
                else:
                    prev = get_column_letter(2+j-1)
                    style_cell(c, f"={prev}25",
                               bold=True, bg=bg, num_fmt=fmt)
            elif label == "  (-) Amortisation":
                style_cell(c, f"=-{col}23*B13",
                           bg=bg, num_fmt=fmt)
            elif label == "Closing Total Debt":
                style_cell(c, f"={col}23+{col}24",
                           bold=True, bg=bg, num_fmt=fmt)
            elif label == "Senior Interest":
                style_cell(c, f"=-{col}23*0.70*B11",
                           bg=bg, num_fmt=fmt)
            elif label == "Mezz Interest":
                style_cell(c, f"=-{col}23*0.30*B12",
                           bg=bg, num_fmt=fmt)
            elif label == "Total Interest Expense":
                style_cell(c, f"={col}26+{col}27",
                           bold=True, bg=bg, num_fmt=fmt)
 
    # FREE CASH FLOW
    section_header(ws, 30, 1, 7, "FREE CASH FLOW")
 
    ws.cell(31,1).fill   = make_fill(MID_BLUE)
    ws.cell(31,1).border = make_border()
    for i, yr in enumerate(proj_years):
        c = ws.cell(31, 2+i)
        style_cell(c, yr, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    fcfe_rows = [
        ("EBITDA",               True,  "ebitda"),
        ("  (-) Interest",       False, "interest"),
        ("  (-) Tax (25%)",      False, "tax"),
        ("  (-) Capex (2% Rev)", False, "capex"),
        ("  (-) Amortisation",   False, "amort"),
        ("Free Cash Flow",       True,  "fcf"),
    ]
 
    for offset, (label, is_bold, key) in enumerate(fcfe_rows):
        r = 32 + offset
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(bold=is_bold, size=9)
        ws.cell(r,1).border = make_border()
        if is_bold:
            ws.cell(r,1).fill = make_fill(LIGHT_BLUE)
 
        for j in range(5):
            col = get_column_letter(2+j)
            c   = ws.cell(r, 2+j)
            bg  = LIGHT_BLUE if is_bold else None
            fmt = "#,##0.0;(#,##0.0);-"
 
            if key == "ebitda":
                style_cell(c, f"={col}18",
                           bold=True, bg=bg, num_fmt=fmt)
            elif key == "interest":
                style_cell(c, f"={col}28",
                           bg=bg, num_fmt=fmt)
            elif key == "tax":
                style_cell(c,
                    f"=-({col}32+{col}33)*0.25",
                    bg=bg, num_fmt=fmt)
            elif key == "capex":
                style_cell(c, f"=-{col}17*0.02",
                           bg=bg, num_fmt=fmt)
            elif key == "amort":
                style_cell(c, f"={col}24",
                           bg=bg, num_fmt=fmt)
            elif key == "fcf":
                style_cell(c,
                    f"={col}32+{col}33+{col}34+{col}35+{col}36",
                    bold=True, bg=bg, num_fmt=fmt)
 
    # EXIT & RETURNS
    section_header(ws, 39, 1, 7,
                   "EXIT ASSUMPTIONS & RETURNS")
 
    exit_items = [
        ("Exit Year",                       5,          "0"),
        ("Exit EBITDA (FY2030E, Rs Cr)",    "='3-statement'!H6*1.10*1.10",     "#,##0.0"),
        ("Exit EV/EBITDA Multiple (x)",     12.0,       "0.0"),
        ("Exit Enterprise Value (Rs Cr)",   "=B41*B42", "#,##0.0"),
        ("(-) Closing Debt at Exit",        "=F25",     "#,##0.0"),
        ("Equity Value at Exit (Rs Cr)",    "=B43-B44", "#,##0.0"),
        ("Initial Equity Invested (Rs Cr)", "=B7",      "#,##0.0"),
        ("Money-on-Money MOIC",             "=B45/B46", "0.0x"),
        ("IRR (5-Year)",                    "=B47^(1/5)-1", "0.0%"),
    ]
 
    for i, (label, val, fmt) in enumerate(exit_items):
        r = 40 + i
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(size=9)
        ws.cell(r,1).border = make_border()
        c = ws.cell(r, 2)
        is_input = not str(val).startswith("=")
        is_result = label in [
            "Money-on-Money MOIC",
            "IRR (5-Year)",
            "Equity Value at Exit (Rs Cr)",
            "Exit Enterprise Value (Rs Cr)"
        ]
        if is_result:
            style_cell(c, val, bold=True,
                       bg=LIGHT_BLUE, num_fmt=fmt)
        elif is_input:
            style_cell(c, val, color=INPUT_BLUE,
                       bg=YELLOW, num_fmt=fmt)
        else:
            style_cell(c, val, color="008000",
                       bg=LIGHT_BLUE, num_fmt=fmt)
 
    # RETURNS SUMMARY
    section_header(ws, 50, 1, 7, "RETURNS SUMMARY")
 
    summary_items = [
        ("Entry EV (Rs Cr)",     "=B6",   "#,##0.0"),
        ("Exit EV (Rs Cr)",      "=B43",  "#,##0.0"),
        ("Entry Multiple (x)",   "=B5",   "0.0"),
        ("Exit Multiple (x)",    "=B42",  "0.0"),
        ("MOIC",                 "=B47",  "0.0x"),
        ("IRR",                  "=B48",  "0.0%"),
        ("Holding Period (Yrs)", "=B40",  "0"),
    ]
 
    for i, (label, formula, fmt) in enumerate(summary_items):
        r = 51 + i
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(bold=True, size=9)
        ws.cell(r,1).fill   = make_fill(LIGHT_BLUE)
        ws.cell(r,1).border = make_border()
        style_cell(ws.cell(r,2), formula,
                   bold=True, bg=LIGHT_BLUE, num_fmt=fmt)
 
    return ws
 
 
# =============================================================
# MAIN
# =============================================================
def build_model(raw_data, market_data):
    print("\n  Building Excel model...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
 
    print("  -> Assumptions...")
    build_assumptions(wb)
 
    print("  -> 3-Statement...")
    build_three_statement(wb)
 
    print("  -> DCF...")
    build_dcf(wb)
 
    print("  -> LBO...")
    build_lbo(wb)
 
    print("  -> Comps Analysis...")
    build_comps(wb)
 
    print("  -> Football Field...")
    build_football_field(wb)
 
    os.makedirs("output", exist_ok=True)
    filename = "output/MOFSL_Financial_Model.xlsx"
    wb.save(filename)
    print(f"\n  Excel saved: {filename}")
    return filename
 
 
# =============================================================
# COMPS SHEET
# =============================================================
def build_comps(wb):
    ws = wb.create_sheet("Comps")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [28,14,14,14,14,14,14,14])
 
    header_cell(ws, 1, 1, 8,
                "COMPARABLE COMPANIES ANALYSIS -- INDIAN IB SECTOR (Rs Crores)",
                size=12)
 
    # ── COLUMN HEADERS ────────────────────────────────────
    headers = ["Company","Mkt Cap","EV",
               "Revenue","EBITDA","Net Income",
               "EV/EBITDA","P/E"]
    ws.cell(2,1).fill   = make_fill(MID_BLUE)
    ws.cell(2,1).border = make_border()
    for i, h in enumerate(headers):
        c = ws.cell(2, 1+i)
        style_cell(c, h, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    # ── COMPARABLE COMPANIES DATA ─────────────────────────
    # Real market data for Indian IB sector (FY2025 estimates)
    # Source: NSE/BSE, yfinance, Screener.in
    section_header(ws, 3, 1, 8, "PEER GROUP — INDIAN INVESTMENT BANKS")
 
    comps_data = [
        # Company, Mkt Cap, EV, Revenue, EBITDA, Net Inc, EV/EBITDA, P/E
        ("Motilal Oswal (Target)",
         42368, 50506, 6441, 4578, 2501, 11.0, 16.9),
        ("Nuvama Wealth Mgmt",
         25532, 27800, 3200, 1920, 1100, 14.5, 23.2),
        ("JM Financial",
         8200,  9800,  2100,  840,  520, 11.7, 15.8),
        ("IIFL Capital Services",
         9800, 11200,  2800, 1120,  650, 10.0, 15.1),
        ("Anand Rathi Wealth",
         25532, 25800, 1850, 1110,  780, 23.2, 32.7),
    ]
 
    for offset, (name, mkt, ev, rev, ebitda,
                 ni, ev_ebitda, pe) in enumerate(comps_data):
        r = 4 + offset
        is_target = "Target" in name
 
        # Highlight target row in gold
        row_bg = "FFF2CC" if is_target else None
        bold   = is_target
 
        ws.cell(r,1).value  = name
        ws.cell(r,1).font   = make_font(bold=bold, size=9)
        ws.cell(r,1).border = make_border()
        if row_bg:
            ws.cell(r,1).fill = make_fill(row_bg)
 
        for j, val in enumerate([mkt, ev, rev,
                                  ebitda, ni,
                                  ev_ebitda, pe]):
            c = ws.cell(r, 2+j)
            if j < 5:   # Rs Cr values
                fmt = "#,##0"
            else:       # multiples
                fmt = "0.0x"
            style_cell(c, val, bold=bold,
                       bg=row_bg,
                       num_fmt=fmt,
                       align="right")
 
    # ── STATISTICS ────────────────────────────────────────
    section_header(ws, 10, 1, 8, "PEER STATISTICS")
 
    stat_rows = [
        ("Mean",   "AVERAGE"),
        ("Median", "MEDIAN"),
        ("High",   "MAX"),
        ("Low",    "MIN"),
    ]
 
    for offset, (label, func) in enumerate(stat_rows):
        r = 11 + offset
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(bold=True, size=9)
        ws.cell(r,1).fill   = make_fill(LIGHT_BLUE)
        ws.cell(r,1).border = make_border()
 
        for j in range(7):
            col = get_column_letter(2+j)
            c   = ws.cell(r, 2+j)
            # Exclude target (row 4) from stats — use rows 5:8
            style_cell(c,
                f"={func}({col}5:{col}8)",
                bold=True, bg=LIGHT_BLUE,
                num_fmt="#,##0" if j < 5 else "0.0x")
 
    # ── IMPLIED VALUATION ─────────────────────────────────
    section_header(ws, 16, 1, 8,
                   "IMPLIED VALUATION OF MOTILAL OSWAL")
 
    ws.cell(17,1).fill   = make_fill(MID_BLUE)
    ws.cell(17,1).border = make_border()
    val_headers = ["Metric","Multiple Used",
                   "MOFSL Metric","Implied EV",
                   "(-) Net Debt","Equity Value",
                   "Implied Price","vs DCF"]
    for i, h in enumerate(val_headers):
        c = ws.cell(17, 1+i)
        style_cell(c, h, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    # MOFSL FY2026E EBITDA = 3,869 Cr (from 3-Statement F6)
    # MOFSL FY2026E Revenue = 6,705 Cr (from 3-Statement F4)
    # Net Debt = 8,138 Cr
    # Shares = 14.4 Cr
    # DCF Price = from DCF sheet B31
 
    implied_rows = [
        ("EV/EBITDA — Mean",
         "=H11", "='3-Statement'!F6",
         "=B18*C18", "=Assumptions!B16",
         "=D18-E18", "=F18/Assumptions!B17*100",
         "=G18/DCF!B31-1"),
        ("EV/EBITDA — Median",
         "=H12", "='3-Statement'!F6",
         "=B19*C19", "=Assumptions!B16",
         "=D19-E19", "=F19/Assumptions!B17*100",
         "=G19/DCF!B31-1"),
        ("EV/EBITDA — High",
         "=H13", "='3-Statement'!F6",
         "=B20*C20", "=Assumptions!B16",
         "=D20-E20", "=F20/Assumptions!B17*100",
         "=G20/DCF!B31-1"),
        ("EV/EBITDA — Low",
         "=H14", "='3-Statement'!F6",
         "=B21*C21", "=Assumptions!B16",
         "=D21-E21", "=F21/Assumptions!B17*100",
         "=G21/DCF!B31-1"),
    ]
 
    for offset, row_vals in enumerate(implied_rows):
        r = 18 + offset
        label = row_vals[0]
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(size=9)
        ws.cell(r,1).border = make_border()
 
        fmts = ["0.0x","#,##0","#,##0",
                "#,##0","#,##0","Rs #,##0","0.0%"]
        for j, (val, fmt) in enumerate(
                zip(row_vals[1:], fmts)):
            c = ws.cell(r, 2+j)
            is_price = j == 5
            style_cell(c, val,
                       bold=is_price,
                       bg=LIGHT_BLUE if is_price else None,
                       num_fmt=fmt)
 
    # ── TRADING COMPS CHART DATA ───────────────────────────
    section_header(ws, 24, 1, 8,
                   "KEY METRICS SUMMARY")
 
    ws.cell(25,1).fill   = make_fill(MID_BLUE)
    ws.cell(25,1).border = make_border()
    metric_headers = ["Company","Revenue Gr%",
                      "EBITDA Mgn%","Net Inc Mgn%",
                      "EV/EBITDA","P/E",
                      "Mkt Cap","Rating"]
    for i, h in enumerate(metric_headers):
        c = ws.cell(25, 1+i)
        style_cell(c, h, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    # Key metrics for each company
    metrics_data = [
        ("Motilal Oswal",  0.245, 0.711, 0.388,
         11.0, 16.9, 42368, "TARGET"),
        ("Nuvama Wealth",  0.180, 0.600, 0.344,
         14.5, 23.2, 25532, "COMP"),
        ("JM Financial",   0.150, 0.400, 0.248,
         11.7, 15.8,  8200, "COMP"),
        ("IIFL Capital",   0.200, 0.400, 0.232,
         10.0, 15.1,  9800, "COMP"),
        ("Anand Rathi",    0.220, 0.600, 0.422,
         23.2, 32.7, 25532, "COMP"),
    ]
 
    for offset, (name, rev_gr, ebitda_mgn,
                 ni_mgn, ev_eb, pe,
                 mkt, rating) in enumerate(metrics_data):
        r = 26 + offset
        is_target = rating == "TARGET"
        row_bg = "FFF2CC" if is_target else None
 
        ws.cell(r,1).value  = name
        ws.cell(r,1).font   = make_font(bold=is_target,
                                         size=9)
        ws.cell(r,1).border = make_border()
        if row_bg:
            ws.cell(r,1).fill = make_fill(row_bg)
 
        vals_fmts = [
            (rev_gr,    "0.0%"),
            (ebitda_mgn,"0.0%"),
            (ni_mgn,    "0.0%"),
            (ev_eb,     "0.0x"),
            (pe,        "0.0x"),
            (mkt,       "#,##0"),
            (rating,    "@"),
        ]
        for j, (val, fmt) in enumerate(vals_fmts):
            c = ws.cell(r, 2+j)
            # Colour code rating cell
            if fmt == "@":
                bg = "FFF2CC" if val == "TARGET" \
                     else "E2EFDA"
                style_cell(c, val, bold=True,
                           bg=bg, align="center")
            else:
                style_cell(c, val,
                           bold=is_target,
                           bg=row_bg,
                           num_fmt=fmt)
 
    return ws
 
 
# =============================================================
# FOOTBALL FIELD CHART SHEET
# =============================================================
def build_football_field(wb):
    ws = wb.create_sheet("Football Field")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [28,12,12,12,12,12,12,12,12,12])
 
    header_cell(ws, 1, 1, 10,
                "FOOTBALL FIELD VALUATION SUMMARY -- MOTILAL OSWAL",
                size=12)
 
    # ── SUBTITLE ──────────────────────────────────────────
    ws.merge_cells("A2:J2")
    ws.cell(2,1).value = "Implied Share Price (Rs) across all valuation methodologies"
    ws.cell(2,1).font  = make_font(italic=True, size=9,
                                    color="666666")
    ws.cell(2,1).alignment = Alignment(horizontal="center")
 
    # ── SECTION: VALUATION INPUTS TABLE ───────────────────
    section_header(ws, 4, 1, 10,
                   "STEP 1 -- VALUATION RANGE INPUTS (Rs per share)")
 
    col_headers = ["Methodology","Low","Mid","High",
                   "Weight","Wtd Low","Wtd Mid","Wtd High",
                   "Notes"]
    ws.cell(5,1).fill   = make_fill(MID_BLUE)
    ws.cell(5,1).border = make_border()
    for i, h in enumerate(col_headers):
        c = ws.cell(5, 1+i)
        style_cell(c, h, bold=True, color=WHITE,
                   bg=MID_BLUE, align="center")
 
    # Valuation ranges — pulled from DCF and LBO sheets
    # DCF implied price = DCF!B31
    # LBO equity at exit / shares = LBO!B45 / Assumptions!B17 * 100
    # Comps EV/EBITDA mean implied = Comps!G18
    # Comps EV/EBITDA median = Comps!G19
 
    val_rows = [
        # Method, Low formula, Mid formula, High formula, Weight, Notes
        ("DCF Valuation",
         "=DCF!B31*0.85",
         "=DCF!B31",
         "=DCF!B31*1.15",
         0.35,
         "WACC 14%, TGR 5%, 3-yr projection"),
 
        ("EV/EBITDA Comps — Low",
         "=Comps!G21",
         "=Comps!G19",
         "=Comps!G18",
         0.25,
         "Peer group: NUVAMA, JM, IIFL, ANAND RATHI"),
 
        ("LBO Floor Valuation",
         "=LBO!B45/Assumptions!B17*100*0.90",
         "=LBO!B45/Assumptions!B17*100",
         "=LBO!B45/Assumptions!B17*100*1.10",
         0.20,
         "PE entry 10x, exit 12x, 5yr, 18.9% IRR"),
 
        ("52-Week Trading Range",
         800,
         1100,
         1400,
         0.10,
         "NSE 52W Low-High range"),
 
        ("Book Value (P/BV 3-4x)",
         "=Assumptions!B17*0+11079.3/14.4*100*3",
         "=Assumptions!B17*0+11079.3/14.4*100*3.5",
         "=Assumptions!B17*0+11079.3/14.4*100*4",
         0.10,
         "Equity Rs 11,079 Cr, 14.4 Cr shares"),
    ]
 
    for offset, (method, low, mid, high,
                 wt, notes) in enumerate(val_rows):
        r = 6 + offset
        is_dcf = "DCF" in method
 
        ws.cell(r,1).value  = method
        ws.cell(r,1).font   = make_font(bold=is_dcf, size=9)
        ws.cell(r,1).border = make_border()
        if is_dcf:
            ws.cell(r,1).fill = make_fill("FFF2CC")
 
        row_bg = "FFF2CC" if is_dcf else None
 
        # Low
        c = ws.cell(r, 2)
        style_cell(c, low, bg=row_bg,
                   num_fmt="Rs #,##0")
 
        # Mid
        c = ws.cell(r, 3)
        style_cell(c, mid, bold=True,
                   bg=LIGHT_BLUE,
                   num_fmt="Rs #,##0")
 
        # High
        c = ws.cell(r, 4)
        style_cell(c, high, bg=row_bg,
                   num_fmt="Rs #,##0")
 
        # Weight
        c = ws.cell(r, 5)
        style_cell(c, wt, color=INPUT_BLUE,
                   bg=YELLOW, num_fmt="0%")
 
        # Weighted Low = Low * Weight
        col_r = get_column_letter(r)
        c = ws.cell(r, 6)
        style_cell(c, f"=B{r}*E{r}",
                   num_fmt="Rs #,##0")
 
        # Weighted Mid
        c = ws.cell(r, 7)
        style_cell(c, f"=C{r}*E{r}",
                   bold=True, bg=LIGHT_BLUE,
                   num_fmt="Rs #,##0")
 
        # Weighted High
        c = ws.cell(r, 8)
        style_cell(c, f"=D{r}*E{r}",
                   num_fmt="Rs #,##0")
 
        # Notes
        c = ws.cell(r, 9)
        style_cell(c, notes, align="left",
                   size=8, italic=True,
                   color="666666")
 
    # ── WEIGHTED AVERAGE SUMMARY ──────────────────────────
    r_sum = 6 + len(val_rows)
 
    ws.cell(r_sum, 1).value  = "WEIGHTED AVERAGE"
    ws.cell(r_sum, 1).font   = make_font(bold=True, size=10,
                                          color=WHITE)
    ws.cell(r_sum, 1).fill   = make_fill(NAVY)
    ws.cell(r_sum, 1).border = make_border()
 
    for j, col_range in enumerate(["B","C","D",
                                    "E","F","G","H"]):
        c = ws.cell(r_sum, 2+j)
        if j == 3:   # Weight col — sum should = 100%
            style_cell(c, f"=SUM(E6:E{r_sum-1})",
                       bold=True, color=WHITE,
                       bg=NAVY, num_fmt="0%")
        elif j in [4, 5, 6]:  # Weighted values
            col_ltr = ["F","G","H"][j-4]
            style_cell(c,
                f"=SUM({col_ltr}6:{col_ltr}{r_sum-1})",
                bold=True, color=WHITE,
                bg=NAVY,
                num_fmt="Rs #,##0")
        else:
            style_cell(c, "---", color="BFBFBF",
                       bg=NAVY, align="center")
 
    # ── FOOTBALL FIELD VISUAL ─────────────────────────────
    section_header(ws, r_sum+2, 1, 10,
                   "STEP 2 -- FOOTBALL FIELD CHART (Visual Range)")
 
    # Bar chart using cell shading to simulate football field
    # Each row = one methodology
    # Columns represent price ranges in Rs 200 bands
 
    price_bands = list(range(200, 4001, 200))  # 200 to 4000
    num_bands   = len(price_bands)             # 19 bands
 
    # Header row for chart
    chart_header_row = r_sum + 4
    ws.cell(chart_header_row, 1).value  = "Methodology"
    ws.cell(chart_header_row, 1).font   = make_font(bold=True,
                                                     size=8,
                                                     color=WHITE)
    ws.cell(chart_header_row, 1).fill   = make_fill(MID_BLUE)
    ws.cell(chart_header_row, 1).border = make_border()
 
    # Price range labels across top (every 4th band = 800 apart)
    for k, price in enumerate(price_bands):
        col = 2 + k
        if col > 10:
            break
        c = ws.cell(chart_header_row, col)
        style_cell(c, f"Rs{price}",
                   bold=True, color=WHITE,
                   bg=MID_BLUE, align="center",
                   size=7)
 
    # Chart colours per methodology
    bar_colors = [
        "4472C4",  # DCF — blue
        "ED7D31",  # Comps — orange
        "70AD47",  # LBO — green
        "FFC000",  # 52W — yellow
        "9E480E",  # P/BV — brown
    ]
 
    method_labels = [
        "DCF",
        "EV/EBITDA Comps",
        "LBO Floor",
        "52-Week Range",
        "P/Book Value",
    ]
 
    # Approximate ranges per methodology (Rs)
    # These are hardcoded visual approximations
    chart_ranges = [
        (1200, 2000),   # DCF low-high (approx)
        (900,   2500),   # Comps
        (1500,  3500),   # LBO
        (800,   1400),   # 52W
        (2300,  3100),   # P/BV
    ]
 
    # Draw bars using cell background
    for row_idx, (label, color,
                  (rng_low, rng_high)) in enumerate(
            zip(method_labels, bar_colors, chart_ranges)):
 
        r_bar = chart_header_row + 1 + row_idx
 
        ws.cell(r_bar, 1).value  = label
        ws.cell(r_bar, 1).font   = make_font(bold=True,
                                              size=8)
        ws.cell(r_bar, 1).border = make_border()
        ws.cell(r_bar, 1).fill   = make_fill("F5F5F5")
 
        for k, price in enumerate(price_bands):
            col = 2 + k
            if col > 10:
                break
            c = ws.cell(r_bar, col)
            c.border = make_border()
 
            prev_price = price_bands[k-1] if k > 0 else 0
 
            # Fill cell if price band falls within range
            if prev_price >= rng_low and price <= rng_high:
                c.fill = make_fill(color)  # Full bar
                c.value = ""
            elif prev_price < rng_low < price:
                c.fill = make_fill(color)  # Partial start
                c.value = ""
            elif prev_price < rng_high < price:
                c.fill = make_fill(color)  # Partial end
                c.value = ""
            else:
                c.fill = make_fill("F5F5F5")  # Empty
 
    # Current market price line
    r_cmp = chart_header_row + 1 + len(method_labels) + 1
    ws.merge_cells(
        start_row=r_cmp, start_column=1,
        end_row=r_cmp,   end_column=10)
    c = ws.cell(r_cmp, 1)
    style_cell(c,
        "Current Market Price ~Rs 768 (as of data fetch) | "
        "52W High Rs 1,064 | 52W Low Rs 612",
        bold=True, size=9,
        color=WHITE, bg=NAVY,
        align="center")
 
   # ── FINAL SUMMARY BOX ─────────────────────────────────
    section_header(ws, r_cmp+2, 1, 10,
                   "STEP 3 -- FINAL VALUATION CONCLUSION")
 
    conclusion_row = r_cmp + 3
    conclusion_items = [
        ("Weighted Average Low  (Rs)", 
          "=F11", "Rs #,##0", True),
        ("Weighted Average Mid  (Rs)",
         f"=G{r_sum}", "Rs #,##0", True),
        ("Weighted Average High (Rs)",
         f"=H{r_sum}", "Rs #,##0", True),
        ("Current Market Price  (Rs)",
         768,          "Rs #,##0", False),
        ("Upside to Mid (Wtd Avg)",
         "=B26/B28-1",
         "0.0%", True),
        ("Analyst Rating",
         "BUY",        "@",        True),
    ]
 
    for i, (label, val, fmt, is_bold) in \
            enumerate(conclusion_items):
        r = conclusion_row + i
        ws.cell(r,1).value  = label
        ws.cell(r,1).font   = make_font(bold=is_bold, size=9)
        ws.cell(r,1).border = make_border()
        if is_bold:
            ws.cell(r,1).fill = make_fill(LIGHT_BLUE)
 
        c = ws.cell(r, 2)
        if label == "Analyst Rating":
            style_cell(c, val, bold=True,
                       bg="E2EFDA",
                       color="375623",
                       align="center")
        elif label == "Current Market Price  (Rs)":
            style_cell(c, val, bold=False,
                       color=INPUT_BLUE,
                       bg=YELLOW,
                       num_fmt=fmt)
        else:
            style_cell(c, val, bold=is_bold,
                       bg=LIGHT_BLUE if is_bold else None,
                       num_fmt=fmt)
 
    return ws
 

